"""
06_descriptive_statistics.py
=============================
Comprehensive descriptive statistics for the merged firm-level panel.

ANGLES COVERED
--------------
  01 - Full distributional statistics for all key variables
       (N, mean, SD, min, p5, p10, p25, median, p75, p90, p95, max, skew, kurtosis)
  02 - Time-series view: variable means and macro series by fiscal year
  03 - Sector-level breakdown: N, mean, median, SD per sector
  04 - Size-quartile breakdown: small / medium-small / medium-large / large
  05 - Debt-status breakdown: zero-interest vs debt-carrying firms
  06 - Panel balance: years-per-firm distribution; entry and exit by year
  07 - Sample-flag composition: how many obs in each estimation sample
  08 - Correlation matrix of core regression variables
  09 - Macro series: one row per year (r_SBP, CPI, delta_r_SBP)
  10 - Sector x Year heatmap of mean CoC_Proxy
  11 - Raw financials summary in PKR billions (TotalAssets, Sales, EBIT, etc.)
  12 - Source-file comparison: old (FY2015-23) vs FY25 extension (FY2024-25)

OUTPUTS
-------
  Results/desc_01_full_variable_stats.csv
  Results/desc_02_by_fiscal_year.csv
  Results/desc_03_by_sector.csv
  Results/desc_04_by_size_quartile.csv
  Results/desc_05_by_debt_status.csv
  Results/desc_06_panel_balance.csv
  Results/desc_07_sample_flags.csv
  Results/desc_08_correlation.csv
  Results/desc_09_macro_series.csv
  Results/desc_10_sector_year_heatmap.csv
  Results/desc_11_raw_financials.csv
  Results/desc_12_source_comparison.csv
  Results/descriptive_statistics_master.xlsx   <- all tables in one workbook

INPUT
-----
  Extracted Data/09_merged_panel.csv

NOTE ON UNITS
-------------
  All monetary items (TotalAssets, Sales, EBIT, InterestExpense, etc.) are
  stored in PKR thousands (000s) in the source workbooks. They are converted
  to PKR billions (000,000,000s) in the raw-financials table for readability.
  Ratio variables (Inv_Rate, CoC_Proxy, P8_ROIC, etc.) are in decimal form.

Reference: Gormsen & Huber (2024, 2025).
"""

import os
import sys
import warnings
import numpy as np
import pandas as pd
from scipy import stats as scipy_stats

warnings.filterwarnings("ignore")

CODES_DIR     = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR   = os.path.dirname(CODES_DIR)
EXTRACTED_DIR = os.path.join(PROJECT_DIR, "Extracted Data")
RESULTS_DIR   = os.path.join(PROJECT_DIR, "Results")

INPUT_PANEL = os.path.join(EXTRACTED_DIR, "09_merged_panel.csv")

os.makedirs(RESULTS_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Sector name normalisation
# The two publication vintages use slightly different sector labels for the
# same underlying groupings. Normalise to a single canonical name for all
# cross-sector statistics.
# ---------------------------------------------------------------------------
SECTOR_NORM = {
    "Textile Sector":     "Textile",
    "Fuel and Energy Sector": "Fuel and Energy",
}

# Short display labels for the heatmap table (sector names are long)
SECTOR_SHORT = {
    "Textile":                                             "Textile",
    "Chemicals, Chemical Products and Pharmaceuticals":    "Chemicals & Pharma",
    "Manufacturing":                                       "Manufacturing",
    "Motor Vehicles, Trailers & Autoparts":                "Auto",
    "Food Products":                                       "Food",
    "Cement":                                              "Cement",
    "Fuel and Energy":                                     "Fuel & Energy",
    "Information and Communication Services":              "ICT",
    "Coke and Refined Petroleum Products":                 "Coke & Petroleum",
    "Other Services Activities":                           "Other Services",
    "Mineral products":                                    "Minerals",
    "Paper, Paperboard and Products":                      "Paper",
    "Sugar":                                               "Sugar",
    "Electrical Machinery and Apparatus":                  "Electrical Machinery",
}

# Core research variables for most tables
CORE_VARS = [
    "Inv_Rate", "CoC_Proxy", "P8_ROIC", "S1_Leverage",
    "SalesGrowth", "Size_ln", "Cashflow", "DepRate",
    "Real_CoC", "delta_r_SBP",
]

# Raw financial items (PKR thousands in data; convert to PKR billions)
RAW_ITEMS = ["TotalAssets", "Sales", "EBIT", "InterestExpense",
             "TotalDebt", "CapEmployed", "Retention", "Depreciation"]

PKR_THOU_TO_BN = 1e-6   # PKR thousands -> PKR billions


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def full_stats(series: pd.Series) -> dict:
    """Return a rich set of distributional statistics for a numeric series."""
    s = series.dropna()
    if len(s) == 0:
        return {}
    sk = float(scipy_stats.skew(s, bias=False))
    ku = float(scipy_stats.kurtosis(s, bias=False))   # excess kurtosis
    return {
        "N":         len(s),
        "Mean":      s.mean(),
        "SD":        s.std(),
        "Min":       s.min(),
        "p5":        s.quantile(0.05),
        "p10":       s.quantile(0.10),
        "p25":       s.quantile(0.25),
        "Median":    s.quantile(0.50),
        "p75":       s.quantile(0.75),
        "p90":       s.quantile(0.90),
        "p95":       s.quantile(0.95),
        "Max":       s.max(),
        "Skewness":  sk,
        "Ex_Kurtosis": ku,
    }


def group_summary(df: pd.DataFrame, groupcol: str,
                  variables: list) -> pd.DataFrame:
    """Mean, median, SD and N for each group x variable."""
    rows = []
    for grp, gdf in df.groupby(groupcol):
        base = {groupcol: grp, "N_obs": len(gdf),
                "N_firms": gdf["firm_id"].nunique()}
        for v in variables:
            if v not in gdf.columns:
                continue
            s = gdf[v].dropna()
            base[f"{v}_mean"]   = s.mean()  if len(s) else np.nan
            base[f"{v}_median"] = s.median() if len(s) else np.nan
            base[f"{v}_sd"]     = s.std()   if len(s) else np.nan
        rows.append(base)
    return pd.DataFrame(rows)


def save(df: pd.DataFrame, fname: str, label: str) -> str:
    """Save a DataFrame to Results/ and print a short summary."""
    path = os.path.join(RESULTS_DIR, fname)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    print(f"  [{label}] -> {fname}  ({len(df)} rows x {df.shape[1]} cols)")
    return path


# ---------------------------------------------------------------------------
# Load and prepare
# ---------------------------------------------------------------------------

def load_panel() -> pd.DataFrame:
    df = pd.read_csv(INPUT_PANEL)
    # Normalise sector labels
    df["sector_norm"] = df["sector"].replace(SECTOR_NORM)
    # Size quartile (based on TotalAssets, computed over full panel)
    df["size_q"] = pd.qcut(
        df["TotalAssets"], q=4,
        labels=["Q1 Small", "Q2 Medium-Small",
                "Q3 Medium-Large", "Q4 Large"]
    )
    print(f"Panel loaded: {len(df):,} rows | {df['firm_id'].nunique():,} firms "
          f"| FY{df['fiscal_year'].min()}-FY{df['fiscal_year'].max()}")
    return df


# ---------------------------------------------------------------------------
# Table 01 — Full distributional statistics
# ---------------------------------------------------------------------------

def table_01_full_stats(df: pd.DataFrame) -> pd.DataFrame:
    """
    Distributional statistics (N through excess kurtosis) for all core
    variables plus the main raw financial items.
    """
    var_labels = {
        "Inv_Rate":         "Investment Rate (delta OFA / lagged Assets)",
        "CoC_Proxy":        "Cost of Capital Proxy (InterestExp / TotalDebt)",
        "P8_ROIC":          "Return on Capital Employed (EBIT / avg CapEmployed)",
        "S1_Leverage":      "Leverage Ratio ((D+E) / Equity)",
        "SalesGrowth":      "Sales Growth Rate",
        "Size_ln":          "Firm Size (ln TotalAssets)",
        "Cashflow":         "Cash Flow Rate (Retention / TotalAssets)",
        "DepRate":          "Depreciation Rate (Depreciation / TotalAssets)",
        "Real_CoC":         "Real Cost of Capital (CoC - CPI/100)",
        "delta_r_SBP":      "Change in SBP Policy Rate (pp)",
        "r_SBP":            "SBP Policy Rate (%)",
        "CPI_inflation":    "CPI Inflation Rate (%)",
        "CoC_Proxy_L1":     "CoC Proxy Lagged 1 Year",
        "CoC_Proxy_L2":     "CoC Proxy Lagged 2 Years",
        "Interaction":      "Interaction (CoC_Proxy x delta_r_SBP)",
    }
    rows = []
    for var, label in var_labels.items():
        if var not in df.columns:
            continue
        st = full_stats(df[var])
        if st:
            st["Variable"] = var
            st["Description"] = label
            rows.append(st)

    out = pd.DataFrame(rows)[
        ["Variable", "Description", "N", "Mean", "SD", "Min",
         "p5", "p10", "p25", "Median", "p75", "p90", "p95",
         "Max", "Skewness", "Ex_Kurtosis"]
    ].round(4)
    return out


# ---------------------------------------------------------------------------
# Table 02 — By fiscal year
# ---------------------------------------------------------------------------

def table_02_by_year(df: pd.DataFrame) -> pd.DataFrame:
    """
    Year-by-year panel of: N firms, N obs, and mean + median of core
    variables, plus the macro values for that year.
    """
    rows = []
    for yr, gdf in df.groupby("fiscal_year"):
        row = {
            "fiscal_year": yr,
            "N_firms":     gdf["firm_id"].nunique(),
            "N_obs":       len(gdf),
            # Macro (year-level constants — take first non-null)
            "r_SBP":        gdf["r_SBP"].iloc[0],
            "CPI_inflation": gdf["CPI_inflation"].iloc[0],
            "delta_r_SBP":   gdf["delta_r_SBP"].iloc[0],
        }
        for v in CORE_VARS:
            if v in ("delta_r_SBP",):
                continue   # already added
            if v not in gdf.columns:
                continue
            s = gdf[v].dropna()
            row[f"{v}_mean"]   = round(s.mean(),   4) if len(s) else np.nan
            row[f"{v}_median"] = round(s.median(), 4) if len(s) else np.nan
            row[f"{v}_sd"]     = round(s.std(),    4) if len(s) else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Table 03 — By sector
# ---------------------------------------------------------------------------

def table_03_by_sector(df: pd.DataFrame) -> pd.DataFrame:
    """
    Per-sector: firm count, obs count, years spanned, and mean/median of
    core variables. Uses normalised sector labels.
    """
    rows = []
    for sec, gdf in df.groupby("sector_norm"):
        yrs = sorted(gdf["fiscal_year"].unique())
        row = {
            "sector":       sec,
            "N_firms":      gdf["firm_id"].nunique(),
            "N_obs":        len(gdf),
            "year_min":     min(yrs),
            "year_max":     max(yrs),
            "N_years_span": len(yrs),
            "pct_zero_int": round(100 * gdf["zero_interest"].mean(), 1)
                            if "zero_interest" in gdf.columns else np.nan,
            "pct_neg_eq":   round(100 * gdf["negative_equity"].mean(), 1)
                            if "negative_equity" in gdf.columns else np.nan,
        }
        for v in CORE_VARS:
            if v in gdf.columns:
                s = gdf[v].dropna()
                row[f"{v}_mean"]   = round(s.mean(),   4) if len(s) else np.nan
                row[f"{v}_median"] = round(s.median(), 4) if len(s) else np.nan
        rows.append(row)

    out = pd.DataFrame(rows).sort_values("N_firms", ascending=False)
    return out


# ---------------------------------------------------------------------------
# Table 04 — By size quartile
# ---------------------------------------------------------------------------

def table_04_by_size(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compare key variables across four size quartiles (Q1=small, Q4=large).
    Also reports median TotalAssets in PKR billions for context.
    """
    rows = []
    for q, gdf in df.groupby("size_q", observed=True):
        row = {
            "size_quartile": str(q),
            "N_firms":       gdf["firm_id"].nunique(),
            "N_obs":         len(gdf),
            "median_assets_PKR_bn": round(
                gdf["TotalAssets"].median() * PKR_THOU_TO_BN, 2),
        }
        for v in CORE_VARS:
            if v in gdf.columns:
                s = gdf[v].dropna()
                row[f"{v}_mean"]   = round(s.mean(),   4) if len(s) else np.nan
                row[f"{v}_median"] = round(s.median(), 4) if len(s) else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Table 05 — By debt status
# ---------------------------------------------------------------------------

def table_05_by_debt(df: pd.DataFrame) -> pd.DataFrame:
    """
    Zero-interest firms vs debt-carrying firms. Zero-interest firms have
    CoC_Proxy = 0 by construction; this table documents how they differ
    on other dimensions to motivate their treatment in robustness checks.
    """
    if "zero_interest" not in df.columns:
        return pd.DataFrame()

    rows = []
    for status, label in [(0, "Debt-carrying (InterestExp > 0)"),
                          (1, "Zero-interest (InterestExp = 0)")]:
        gdf = df[df["zero_interest"] == status]
        row = {
            "debt_status": label,
            "N_firms":     gdf["firm_id"].nunique(),
            "N_obs":       len(gdf),
            "pct_of_panel": round(100 * len(gdf) / len(df), 1),
            "median_assets_PKR_bn": round(
                gdf["TotalAssets"].median() * PKR_THOU_TO_BN, 2)
                if len(gdf) else np.nan,
        }
        for v in ["Inv_Rate", "P8_ROIC", "S1_Leverage", "SalesGrowth",
                  "Size_ln", "Cashflow"]:
            if v in gdf.columns:
                s = gdf[v].dropna()
                row[f"{v}_mean"]   = round(s.mean(),   4) if len(s) else np.nan
                row[f"{v}_median"] = round(s.median(), 4) if len(s) else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Table 06 — Panel balance
# ---------------------------------------------------------------------------

def table_06_panel_balance(df: pd.DataFrame) -> tuple:
    """
    Returns:
      balance_df  : per-firm years-in-panel count
      entry_exit  : per-year entry / exit / active firm counts
    """
    # Years per firm
    yrs_per_firm = df.groupby("firm_id")["fiscal_year"].agg(
        ["count", "min", "max"]
    ).rename(columns={"count": "N_years", "min": "first_year", "max": "last_year"})
    yrs_per_firm["N_years"] = yrs_per_firm["N_years"].astype(int)

    balance_summary = (
        yrs_per_firm["N_years"].value_counts().sort_index()
        .reset_index().rename(columns={"N_years": "years_in_panel",
                                       "count": "N_firms"})
    )
    balance_summary["pct_firms"] = (
        100 * balance_summary["N_firms"] / balance_summary["N_firms"].sum()
    ).round(1)

    # Entry / exit by year
    all_years = sorted(df["fiscal_year"].unique())
    entry_exit_rows = []
    for yr in all_years:
        active  = set(df[df["fiscal_year"] == yr]["firm_id"])
        if yr > min(all_years):
            prev    = set(df[df["fiscal_year"] == yr - 1]["firm_id"])
            entered = len(active - prev)
            exited  = len(prev - active)
        else:
            entered = len(active)
            exited  = 0
        entry_exit_rows.append({
            "fiscal_year": yr,
            "N_active": len(active),
            "N_entered": entered,
            "N_exited": exited,
        })
    entry_exit = pd.DataFrame(entry_exit_rows)
    return balance_summary, entry_exit


# ---------------------------------------------------------------------------
# Table 07 — Sample flag composition
# ---------------------------------------------------------------------------

def table_07_flags(df: pd.DataFrame) -> pd.DataFrame:
    """Count and percentage for each sample flag and key intersections."""
    flags = [
        "in_model1_sample", "in_model3_sample", "balanced_subsample",
        "zero_interest", "negative_equity", "non_june_fye",
        "robustness_excl", "post2023",
    ]
    rows = []
    total = len(df)
    for f in flags:
        if f not in df.columns:
            continue
        n    = int(df[f].sum())
        n_no = total - n
        rows.append({
            "flag":         f,
            "N_flag_eq_1":  n,
            "N_flag_eq_0":  n_no,
            "pct_eq_1":     round(100 * n / total, 1),
            "N_firms":      df[df[f] == 1]["firm_id"].nunique() if n else 0,
        })

    # Key intersections for regression readiness
    inter_rows = [
        ("Model1 & balanced",
         df[(df.get("in_model1_sample", 0) == 1) &
            (df.get("balanced_subsample", 0) == 1)]),
        ("Model3 & balanced",
         df[(df.get("in_model3_sample", 0) == 1) &
            (df.get("balanced_subsample", 0) == 1)]),
        ("Model1 & debt-carrying",
         df[(df.get("in_model1_sample", 0) == 1) &
            (df.get("zero_interest", 1) == 0)]),
        ("Model1 & no robustness_excl",
         df[(df.get("in_model1_sample", 0) == 1) &
            (df.get("robustness_excl", 1) == 0)]),
    ]
    for label, sub in inter_rows:
        rows.append({
            "flag":        label,
            "N_flag_eq_1": len(sub),
            "N_flag_eq_0": total - len(sub),
            "pct_eq_1":    round(100 * len(sub) / total, 1),
            "N_firms":     sub["firm_id"].nunique(),
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Table 08 — Correlation matrix
# ---------------------------------------------------------------------------

def table_08_correlation(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pearson correlation matrix for core regression variables.
    Computed on observations where all variables are non-null (pairwise).
    """
    corr_vars = [
        "Inv_Rate", "CoC_Proxy", "CoC_Proxy_L1", "CoC_Proxy_L2",
        "P8_ROIC", "S1_Leverage", "SalesGrowth", "Size_ln",
        "Cashflow", "Real_CoC", "delta_r_SBP",
    ]
    avail = [v for v in corr_vars if v in df.columns]
    corr  = df[avail].corr(method="pearson").round(3)
    corr.index.name = "Variable"
    return corr.reset_index()


# ---------------------------------------------------------------------------
# Table 09 — Macro series
# ---------------------------------------------------------------------------

def table_09_macro(df: pd.DataFrame) -> pd.DataFrame:
    """One row per fiscal year with the macro variables and panel size."""
    rows = []
    for yr in sorted(df["fiscal_year"].unique()):
        gdf = df[df["fiscal_year"] == yr]
        rows.append({
            "fiscal_year":  yr,
            "N_firms":      gdf["firm_id"].nunique(),
            "r_SBP_pct":    round(gdf["r_SBP"].iloc[0], 4),
            "CPI_pct":      round(gdf["CPI_inflation"].iloc[0], 2),
            "delta_r_SBP":  round(gdf["delta_r_SBP"].iloc[0], 4),
            "real_rate_pct": round(gdf["r_SBP"].iloc[0] - gdf["CPI_inflation"].iloc[0], 2),
            "mean_CoC":     round(gdf["CoC_Proxy"].mean(), 4),
            "median_CoC":   round(gdf["CoC_Proxy"].median(), 4),
            "mean_Inv":     round(gdf["Inv_Rate"].mean(), 4),
            "median_Inv":   round(gdf["Inv_Rate"].median(), 4),
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Table 10 — Sector x Year heatmap of CoC_Proxy
# ---------------------------------------------------------------------------

def table_10_sector_year(df: pd.DataFrame) -> pd.DataFrame:
    """
    Mean CoC_Proxy (and mean Inv_Rate) by sector (rows) x fiscal year (cols).
    Sectors with fewer than 3 obs in a cell are shown as NaN.
    """
    df2 = df.copy()
    df2["sector_short"] = df2["sector_norm"].map(SECTOR_SHORT).fillna(df2["sector_norm"])

    pivot_coc = df2.pivot_table(
        values="CoC_Proxy", index="sector_short",
        columns="fiscal_year", aggfunc="mean",
        observed=True
    ).round(4)

    # Mask cells with fewer than 3 observations
    count_pivot = df2.pivot_table(
        values="CoC_Proxy", index="sector_short",
        columns="fiscal_year", aggfunc="count",
        observed=True
    )
    pivot_coc[count_pivot < 3] = np.nan

    pivot_coc.columns.name = None
    pivot_coc.index.name   = "Sector"
    pivot_coc["Row_mean"]  = pivot_coc.mean(axis=1).round(4)
    pivot_coc = pivot_coc.sort_values("Row_mean", ascending=False)
    return pivot_coc.reset_index()


# ---------------------------------------------------------------------------
# Table 11 — Raw financials summary
# ---------------------------------------------------------------------------

def table_11_raw_financials(df: pd.DataFrame) -> pd.DataFrame:
    """
    Summary statistics for raw PKR monetary items, expressed in PKR billions.
    Includes mean, median, and IQR to convey the economic scale of firms.
    """
    rows = []
    labels = {
        "TotalAssets":      "Total Assets (PKR bn)",
        "Sales":            "Sales / Revenue (PKR bn)",
        "EBIT":             "EBIT (PKR bn)",
        "InterestExpense":  "Interest Expense (PKR bn)",
        "TotalDebt":        "Total Debt D+E (PKR bn)",
        "CapEmployed":      "Capital Employed (PKR bn)",
        "Retention":        "Retention in Business (PKR bn)",
        "Depreciation":     "Depreciation (PKR bn)",
    }
    for col, label in labels.items():
        if col not in df.columns:
            continue
        s = df[col].dropna() * PKR_THOU_TO_BN
        rows.append({
            "Item":   label,
            "N":      len(s),
            "Mean":   round(s.mean(),   2),
            "SD":     round(s.std(),    2),
            "Min":    round(s.min(),    2),
            "p25":    round(s.quantile(0.25), 2),
            "Median": round(s.median(), 2),
            "p75":    round(s.quantile(0.75), 2),
            "Max":    round(s.max(),    2),
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Table 12 — Source-file comparison
# ---------------------------------------------------------------------------

def table_12_source_comparison(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compare the old-file contribution (FY2015-2023) with the FY25 extension
    (FY2024-2025) across all core variables to document any level shift.
    """
    rows = []
    for src, label in [("old",  "Old file (FY2015-2023)"),
                       ("fy25", "FY25 extension (FY2024-2025)")]:
        gdf = df[df["source_file"] == src]
        row = {
            "source":  label,
            "N_obs":   len(gdf),
            "N_firms": gdf["firm_id"].nunique(),
            "years":   f"{int(gdf['fiscal_year'].min())}-"
                       f"{int(gdf['fiscal_year'].max())}",
        }
        for v in CORE_VARS + ["r_SBP", "CPI_inflation"]:
            if v in gdf.columns:
                s = gdf[v].dropna()
                row[f"{v}_mean"]   = round(s.mean(),   4) if len(s) else np.nan
                row[f"{v}_median"] = round(s.median(), 4) if len(s) else np.nan
        rows.append(row)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Excel workbook writer
# ---------------------------------------------------------------------------

def write_excel(tables: dict, path: str) -> None:
    """
    Write all tables to a single Excel workbook, one sheet per table.
    Applies basic formatting: bold headers, auto column widths, freeze pane.
    If the target file is locked (open in Excel), saves to a timestamped copy.
    """
    from datetime import datetime
    try:
        open(path, "ab").close()   # quick write-permission test
    except PermissionError:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = os.path.splitext(path)
        path = f"{base[0]}_{ts}{base[1]}"
        print(f"  NOTE: original file locked -- saving to {os.path.basename(path)}")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # Bold header row
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            for cell in ws[1]:
                cell.font    = header_fill and header_font
                cell.fill    = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Auto-width (capped at 40 characters)
            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0)
                    for c in col
                )
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            # Freeze first row
            ws.freeze_panes = "A2"

    print(f"\n  Excel workbook -> {os.path.basename(path)}"
          f"  ({len(tables)} sheets)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run():
    print("=" * 65)
    print("Script 06 -- Descriptive Statistics")
    print("=" * 65)

    df = load_panel()
    print()

    saved_files = {}

    # -- Table 01 -----------------------------------------------------------
    print("Building tables...")
    t01 = table_01_full_stats(df)
    save(t01, "desc_01_full_variable_stats.csv", "01")
    saved_files["01_Variable_Stats"] = t01

    # -- Table 02 -----------------------------------------------------------
    t02 = table_02_by_year(df)
    save(t02, "desc_02_by_fiscal_year.csv", "02")
    saved_files["02_By_Year"] = t02

    # -- Table 03 -----------------------------------------------------------
    t03 = table_03_by_sector(df)
    save(t03, "desc_03_by_sector.csv", "03")
    saved_files["03_By_Sector"] = t03

    # -- Table 04 -----------------------------------------------------------
    t04 = table_04_by_size(df)
    save(t04, "desc_04_by_size_quartile.csv", "04")
    saved_files["04_By_Size_Quartile"] = t04

    # -- Table 05 -----------------------------------------------------------
    t05 = table_05_by_debt(df)
    save(t05, "desc_05_by_debt_status.csv", "05")
    saved_files["05_By_Debt_Status"] = t05

    # -- Table 06 -----------------------------------------------------------
    t06_balance, t06_entry = table_06_panel_balance(df)
    save(t06_balance, "desc_06a_panel_balance.csv", "06a")
    save(t06_entry,   "desc_06b_entry_exit.csv",    "06b")
    saved_files["06a_Panel_Balance"]  = t06_balance
    saved_files["06b_Entry_Exit"]     = t06_entry

    # -- Table 07 -----------------------------------------------------------
    t07 = table_07_flags(df)
    save(t07, "desc_07_sample_flags.csv", "07")
    saved_files["07_Sample_Flags"] = t07

    # -- Table 08 -----------------------------------------------------------
    t08 = table_08_correlation(df)
    save(t08, "desc_08_correlation.csv", "08")
    saved_files["08_Correlation"] = t08

    # -- Table 09 -----------------------------------------------------------
    t09 = table_09_macro(df)
    save(t09, "desc_09_macro_series.csv", "09")
    saved_files["09_Macro_Series"] = t09

    # -- Table 10 -----------------------------------------------------------
    t10 = table_10_sector_year(df)
    save(t10, "desc_10_sector_year_CoC_heatmap.csv", "10")
    saved_files["10_SectorYear_CoC"] = t10

    # -- Table 11 -----------------------------------------------------------
    t11 = table_11_raw_financials(df)
    save(t11, "desc_11_raw_financials_PKRbn.csv", "11")
    saved_files["11_Raw_Financials"] = t11

    # -- Table 12 -----------------------------------------------------------
    t12 = table_12_source_comparison(df)
    save(t12, "desc_12_source_comparison.csv", "12")
    saved_files["12_Source_Comparison"] = t12

    # -- Excel workbook -----------------------------------------------------
    xlsx_path = os.path.join(RESULTS_DIR, "descriptive_statistics_master.xlsx")
    write_excel(saved_files, xlsx_path)

    # -- Console summary report ---------------------------------------------
    print()
    print("=" * 65)
    print("  DESCRIPTIVE STATISTICS — CONSOLE SUMMARY")
    print("=" * 65)

    print(f"\n  Panel: {len(df):,} obs | {df['firm_id'].nunique():,} firms "
          f"| FY{df['fiscal_year'].min()}-FY{df['fiscal_year'].max()}")
    print(f"  Sectors: {df['sector_norm'].nunique()} (normalised labels)")
    print(f"  Source split: "
          f"old={( df['source_file']=='old').sum():,}  "
          f"fy25={(df['source_file']=='fy25').sum():,}")

    print("\n  -- Key variable means and medians (full panel) --")
    print(f"  {'Variable':<20s}  {'N':>5s}  {'Mean':>8s}  {'Median':>8s}  {'SD':>8s}")
    print(f"  {'-'*20}  {'-'*5}  {'-'*8}  {'-'*8}  {'-'*8}")
    for _, row in t01.iterrows():
        print(f"  {row['Variable']:<20s}  {int(row['N']):>5,}  "
              f"{row['Mean']:>8.4f}  {row['Median']:>8.4f}  {row['SD']:>8.4f}")

    print("\n  -- Sector composition (by firm count) --")
    print(f"  {'Sector':<50s}  {'Firms':>6s}  {'Obs':>6s}  {'mean_CoC':>9s}")
    print(f"  {'-'*50}  {'-'*6}  {'-'*6}  {'-'*9}")
    for _, row in t03.iterrows():
        coc = row.get("CoC_Proxy_mean", np.nan)
        coc_s = f"{coc:.4f}" if pd.notna(coc) else "   n/a"
        print(f"  {str(row['sector']):<50s}  {int(row['N_firms']):>6,}  "
              f"{int(row['N_obs']):>6,}  {coc_s:>9s}")

    print("\n  -- Size quartile profile --")
    print(f"  {'Quartile':<20s}  {'Firms':>6s}  {'Med Assets PKR bn':>18s}  "
          f"{'mean CoC':>9s}  {'mean Inv':>9s}")
    print(f"  {'-'*20}  {'-'*6}  {'-'*18}  {'-'*9}  {'-'*9}")
    for _, row in t04.iterrows():
        coc = row.get("CoC_Proxy_mean", np.nan)
        inv = row.get("Inv_Rate_mean", np.nan)
        print(f"  {str(row['size_quartile']):<20s}  "
              f"{int(row['N_firms']):>6,}  "
              f"{row['median_assets_PKR_bn']:>18,.1f}  "
              f"{coc:>9.4f}  {inv:>9.4f}")

    print("\n  -- Year-by-year macro and CoC means --")
    print(f"  {'Year':>4s}  {'Firms':>5s}  {'r_SBP':>6s}  {'CPI':>6s}  "
          f"{'d_r':>6s}  {'mean_CoC':>9s}  {'med_CoC':>9s}  {'mean_Inv':>9s}")
    print(f"  {'-'*4}  {'-'*5}  {'-'*6}  {'-'*6}  "
          f"{'-'*6}  {'-'*9}  {'-'*9}  {'-'*9}")
    for _, row in t09.iterrows():
        print(f"  {int(row['fiscal_year']):>4d}  "
              f"{int(row['N_firms']):>5d}  "
              f"{row['r_SBP_pct']:>6.2f}  "
              f"{row['CPI_pct']:>6.2f}  "
              f"{row['delta_r_SBP']:>+6.2f}  "
              f"{row['mean_CoC']:>9.4f}  "
              f"{row['median_CoC']:>9.4f}  "
              f"{row['mean_Inv']:>9.4f}")

    print("\n  -- Panel balance distribution --")
    print(f"  {'Years in panel':>14s}  {'N firms':>8s}  {'%':>6s}")
    for _, row in t06_balance.iterrows():
        print(f"  {int(row['years_in_panel']):>14d}  "
              f"{int(row['N_firms']):>8,}  "
              f"{row['pct_firms']:>5.1f}%")

    print()
    print("=" * 65)
    print(f"All files saved to: {RESULTS_DIR}")
    print("=" * 65)

    return df


if __name__ == "__main__":
    run()
